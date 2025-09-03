import os
import random
import re
from datetime import datetime
from io import BytesIO

import mysql.connector
from flask import Flask, request, render_template, send_file, url_for, flash, session, redirect, g
from flask_mail import Mail, Message
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd


app = Flask(__name__)
app.secret_key = 'your_super_secret_key'

# ✅ Mail Configuration
app.config.update(
    MAIL_SERVER='smtp.gmail.com',
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME='reshmamohan938@gmail.com',
    MAIL_PASSWORD='bbds blto hdeo kjun',
    MAIL_DEFAULT_SENDER ='reshmamohan938@gmail.com'
)
mail = Mail(app)

app = Flask(__name__)

# ✅ MySQL Configuration (manual, mimicking flask_mysqldb)
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'root'
app.config['MYSQL_DB'] = 'softnis_db'

# ✅ Email Configuration
app.config['MAIL_DEFAULT_SENDER'] = 'reshmamohan938@gmail.com'

# ✅ MySQL wrapper class
class MySQLWrapper:
    def __init__(self, app):
        self.config = {
            'host': app.config.get('MYSQL_HOST'),
            'user': app.config.get('MYSQL_USER'),
            'password': app.config.get('MYSQL_PASSWORD'),
            'database': app.config.get('MYSQL_DB')
        }

    @property
    def connection(self):
        if 'db_conn' not in g:
            g.db_conn = mysql.connector.connect(**self.config)
        return g.db_conn

    def close_connection(self, e=None):
        db_conn = g.pop('db_conn', None)
        if db_conn is not None:
            db_conn.close()

mysql = MySQLWrapper(app)
app.teardown_appcontext(mysql.close_connection)

# ✅ Folder setup for uploads/results
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)


@app.route('/')
def start_page():
    return render_template('start_page.html')

@app.route('/admin_email', methods=['GET', 'POST'])
def admin_email():
    allowed_admins = ['reshmamohan938@gmail.com', 'indra@softnis.com']  # ✅ Replace with real emails

    if request.method == 'POST':
        email = request.form['email']

        # Check if the email is one of the allowed admins
        if email not in allowed_admins:
            flash("❌ You are not authorized as an admin.", "popup")
            return render_template('admin_email.html')

        # Generate OTP and store in session
        otp = str(random.randint(1000, 9999))
        session['otp'] = otp
        session['email'] = email
        session['role'] = 'admin'  # Optional for later logic

        msg = Message('Your Admin OTP Login Code',
                      sender='your_email@gmail.com',
                      recipients=[email])
        msg.body = f'Your OTP is: {otp}'
        mail.send(msg)

        flash("✅ OTP sent to your email.", "success")
        return redirect(url_for('verify'))  # or redirect to admin_verify if needed

    return render_template('admin_email.html')


@app.route('/add_employee', methods=['GET', 'POST'])
def add_employee():
    if session.get('role') != 'admin':
        flash("❌ Unauthorized access.", "popup")
        return redirect(url_for('project_options'))

    if request.method == 'POST':
        name = request.form['employee_name']
        email = request.form['employee_email']
        phone = request.form['employee_phone']
        added_by = session.get('email')  # Store which admin added

        cursor = mysql.connection.cursor()
        cursor.execute(
            "INSERT INTO employees (emp_name, email, phone, added_by) VALUES (%s, %s, %s, %s)",
            (name, email, phone, added_by)
        )
        mysql.connection.commit()
        cursor.close()

        flash("✅ Employee added successfully!", "popup")
        return redirect(url_for('add_employee'))

    return render_template('add_employee.html')


@app.route('/add_manager', methods=['GET', 'POST'])
def add_manager():
    if session.get('role') != 'admin':
        flash("❌ Unauthorized access.", "popup")
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        name = request.form['manager_name']
        email = request.form['manager_email']
        phone = request.form['manager_phone']
        cursor = mysql.connection.cursor()
        cursor.execute("INSERT INTO managers (name, email, phone) VALUES (%s, %s, %s)", (name, email, phone))
        mysql.connection.commit()
        flash("✅ Manager added.", "popup")
        return redirect(url_for('add_manager'))

    return render_template("add_manager.html")

@app.route('/view_employees')
def view_employees():
    if session.get('role') != 'admin':
        flash("❌ Unauthorized access.", "popup")
        return redirect(url_for('project_options'))

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM employees")
    employees = cursor.fetchall()

    return render_template("view_employees.html", employees=employees)

@app.route('/delete_employee/<int:emp_id>', methods=['POST'])
def delete_employee(emp_id):
    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM employees WHERE emp_id = %s", (emp_id,))
    mysql.connection.commit()
    cursor.close()
    flash("✅ Employee deleted successfully.", "success")
    return redirect(url_for('view_employees'))


@app.route('/view_managers')
def view_managers():
    if session.get('role') != 'admin':
        flash("❌ Unauthorized access.", "popup")
        return redirect(url_for('project_options'))

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT * FROM managers")  # assuming your table is named `managers`
    managers = cursor.fetchall()
    cursor.close()

    return render_template('view_managers.html', managers=managers)


@app.route('/delete_manager/<int:mang_id>', methods=['POST'])
def delete_manager(mang_id):
    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM managers WHERE mang_id = %s", (mang_id,))
    mysql.connection.commit()
    flash("✅ Manager deleted successfully.", "success")
    return redirect(url_for('view_managers'))


@app.route('/admin_dashboard')   # after admin login there need of 2 option so this is created
def admin_dashboard():
    if session.get('role') != 'admin':
        flash("❌ Unauthorized access.", "popup")
        return redirect(url_for('login'))  # or any other fallback

    return render_template("admin_dashboard.html")


@app.route('/welcome')
def welcome():
    return render_template('welcome.html')

@app.route('/prerequisites')
def prerequisites():
    return render_template('prerequisites.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    allowed_domains = ['@softnis.com']
    admin_emails = ['indra@softnis.com']  # ✅ Admin emails

    if request.method == 'POST':
        email = request.form['email'].strip().lower()

        if not any(email.endswith(domain) for domain in allowed_domains):
            flash("❌ Only emails ending in @softnis.com are accepted.", "popup")
            return render_template('login.html')

        if email in admin_emails:
            session['is_admin'] = True
            session['role'] = 'admin'      # ✅ Set role for admin
        else:
            session['is_admin'] = False
            session['role'] = 'manager'    # ✅ Set role for manager/user

        session['email'] = email

        # ✅ Send OTP
        otp = str(random.randint(1000, 9999))
        session['otp'] = otp
        msg = Message('Your OTP for Login', sender='your_email@gmail.com', recipients=[email])
        msg.body = f'Your OTP is: {otp}'
        mail.send(msg)

        flash("✅ OTP sent to your email.", "success")
        return redirect(url_for('verify'))

    return render_template('login.html')






@app.route('/verify', methods=['GET', 'POST'])
def verify():
    if request.method == 'POST':
        entered_otp = request.form.get('otp')
        actual_otp = session.get('otp')

        if entered_otp == actual_otp:
            session['logged_in'] = True
            role = session.get('role', 'user')

            print("✅ Logged in as:", role)

            flash("✅ Login successful!", "success")
            role = session.get('role', 'user')
            print("DEBUG session['role']:", session.get('role'))

            if role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('project_options'))
        else:
            flash("❌ Incorrect OTP. Please try again.", "popup")
            return render_template('verify.html')

    return render_template('verify.html')



@app.route('/project_options', methods=['GET', 'POST'])
def project_options():
    return render_template('project_options.html')


@app.route('/add_project', methods=['GET', 'POST'])
def add_project():
    if request.method == 'POST':
        project_name = request.form['project_name']
        manager_name = request.form['project_manager']  # manager name as string
        selected_employees = request.form.getlist('selected_employees')

        members = []
        for emp in selected_employees:
            name, email, phone = emp.split("|")
            members.append({
                'name': name.strip(),
                'email': email.strip(),
                'phone': phone.strip()
            })

        try:
            cursor = mysql.connection.cursor()

            # ✅ Do NOT insert project_id (auto_increment)
            cursor.execute(
                "INSERT INTO projects (name, manager) VALUES (%s, %s)",
                (project_name, manager_name)
            )

            # ✅ Get the auto-generated project_id
            project_id = cursor.lastrowid

            # Insert members linked to this project
            for m in members:
                cursor.execute(
                    "INSERT INTO members (project_id, name, email, phone) VALUES (%s, %s, %s, %s)",
                    (project_id, m['name'], m['email'], m['phone'])
                )

            mysql.connection.commit()
            cursor.close()

            flash("✅ Project and members saved successfully!", "popup")
            return redirect(url_for('add_project'))

        except Exception as e:
            print("❌ ERROR:", e)
            flash(f"❌ Database error: {str(e)}", "popup")
            return redirect(url_for('add_project'))

    # GET request: fetch employees and managers
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT emp_name, email, phone FROM employees")
    employees = cursor.fetchall()

    cursor.execute("SELECT name, email, phone FROM managers")
    managers = cursor.fetchall()
    cursor.close()

    return render_template('add_project.html', employees=employees, managers=managers)

@app.route('/add_member_to_session', methods=['POST'])
def add_member_to_session():
    selected_email = request.form['selected_employee_email']

    cursor = mysql.connection.cursor()
    cursor.execute("SELECT emp_name, email, phone FROM employees WHERE email = %s", (selected_email,))
    emp = cursor.fetchone()
    cursor.close()

    if emp:
        member = {
            'name': emp[0],
            'email': emp[1],
            'phone': emp[2]
        }

        members = session.get('members', [])
        if not any(m['email'] == member['email'] for m in members):  # avoid duplicates
            members.append(member)
            session['members'] = members
            flash("✅ Member added to project.", "popup")
        else:
            flash("⚠️ Member already added.", "popup")
    else:
        flash("❌ Employee not found.", "popup")

    return redirect(url_for('add_project'))



@app.route('/assign_members/<int:project_id>', methods=['GET', 'POST'])
def assign_members(project_id):
    if session.get('role') != 'admin':
        flash("❌ Unauthorized access.", "popup")
        return redirect(url_for('project_options'))

    cursor = mysql.connection.cursor()

    if request.method == 'POST':
        selected_employee_ids = request.form.getlist('selected_employees')

        for emp_id in selected_employee_ids:
            cursor.execute("SELECT name, email, phone FROM employees WHERE id = %s", (emp_id,))
            emp = cursor.fetchone()
            if emp:
                cursor.execute("""
                    INSERT INTO members (project_id, name, email, phone)
                    VALUES (%s, %s, %s, %s)
                """, (project_id, emp[0], emp[1], emp[2]))

        mysql.connection.commit()
        cursor.close()
        flash("✅ Members assigned to project!", "success")
        return redirect(url_for('project_options'))

    # GET: fetch all employees
    cursor.execute("SELECT id, name, email FROM employees")
    employees = cursor.fetchall()
    cursor.close()

    return render_template('assign_members.html', employees=employees, project_id=project_id)



@app.route('/existing_projects')
def existing_projects():
    cursor = mysql.connection.cursor()

    search_name = request.args.get('search_name')

    if search_name:
        cursor.execute("""
            SELECT p.project_id, p.name, p.manager, m.name AS manager_name
            FROM projects p
            LEFT JOIN managers m ON p.manager = m.email
            WHERE LOWER(p.name) LIKE LOWER(%s)
        """, (f"%{search_name}%",))
    else:
        cursor.execute("""
            SELECT p.project_id, p.name, p.manager, m.name AS manager_name
            FROM projects p
            LEFT JOIN managers m ON p.manager = m.email
        """)

    projects = cursor.fetchall()

    # Members
    cursor.execute("SELECT project_id, name, email, phone FROM members")
    member_rows = cursor.fetchall()
    members_by_project = {}
    for project_id, name, email, phone in member_rows:
        members_by_project.setdefault(project_id, []).append({
            'name': name,
            'email': email,
            'phone': phone
        })

    # Files
    project_files = {}
    for proj in projects:
        proj_id = proj[0]
        cursor.execute("SELECT filename, filepath, pro_id FROM project_files WHERE project_id = %s", (proj_id,))
        project_files[proj_id] = cursor.fetchall()

    cursor.close()

    return render_template(
        'existing_projects.html',
        projects=projects,
        members_by_project=members_by_project,
        files=project_files
    )



def compare_user_quality(df1, df2, file_name):
    merged = pd.merge(df1, df2, on='SoftNis ID', suffixes=('_Prod', '_Deliv'))

    summary_rows = []
    mismatch_rows = []

    if 'User Name_Prod' not in merged.columns:
        merged['User Name'] = 'Unknown'
    else:
        merged['User Name'] = merged['User Name_Prod']

    users = merged['User Name'].unique()

    for user in users:
        user_data = merged[merged['User Name'] == user]
        total = correct = incorrect = 0

        for _, row in user_data.iterrows():
            for col in df1.columns:
                if col in ['SoftNis ID', 'User Name']:
                    continue
                val1 = row.get(f"{col}_Prod")
                val2 = row.get(f"{col}_Deliv")

                if pd.isna(val1) and pd.isna(val2):
                    continue

                total += 1
                if val1 == val2:
                    correct += 1
                else:
                    incorrect += 1
                    mismatch_rows.append({
                        'Source File': file_name,
                        'SoftNis ID': row['SoftNis ID'],
                        'User Name': user,
                        'Attribute': col,
                        'Production Value': val1,
                        'Delivered Value': val2
                    })

        accuracy = round((correct / total) * 100, 2) if total else 0

        summary_rows.append({
            'User Name': user,
            'Total Fields Checked': total,
            'Correct Fields': correct,
            'Incorrect Fields': incorrect,
            'Accuracy (%)': accuracy,
            'Source File': file_name
        })

    return pd.DataFrame(summary_rows), pd.DataFrame(mismatch_rows)



def get_files_by_project_id(project_id):
    cursor = mysql.connection.cursor()

    cursor.execute("SELECT filename, filepath FROM project_files WHERE project_id = %s", (project_id,))
    result = cursor.fetchall()
    cursor.close()
    return result


def process_quality_reports(file_paths):
    all_data = []

    for file in file_paths:
        print(f"\n📁 Checking file: {file}")

        if not os.path.exists(file):
            print(f"❌ File not found: {file}")
            continue

        try:
            xls = pd.ExcelFile(file)
            print("📄 Available sheets:", xls.sheet_names)

            if "Quality Report" not in xls.sheet_names:
                print(f"⚠️ 'Quality Report' sheet not found in {file}")
                continue

            df = pd.read_excel(xls, sheet_name="Quality Report")
            required_columns = {"User Name", "Matched Cells", "Total Cells", "Quality %"}

            if not required_columns.issubset(df.columns):
                missing = required_columns - set(df.columns)
                print(f"⚠️ Missing columns in {file}: {missing}")
                continue

            df["Source File"] = os.path.basename(file)
            all_data.append(df)
            print(f"✅ Processed successfully: {file}")

        except Exception as e:
            print(f"❌ Error reading {file}: {e}")

    if not all_data:
        print("\n🚫 No valid 'Quality Report' data found in any file.")
        return None, None, []

    # Combine all data
    combined_df = pd.concat(all_data)

    # Create summary
    summary_df = combined_df.groupby("User Name").agg({
        "Matched Cells": "sum",
        "Total Cells": "sum"
    }).reset_index()
    summary_df["Quality %"] = (summary_df["Matched Cells"] / summary_df["Total Cells"]) * 100

    # Prepare graph data for frontend (optional)
    graph_data = summary_df.to_dict(orient='records')

    # File paths
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    summary_path = f"static/summary_report_{timestamp}.xlsx"
    detailed_path = f"static/detailed_report_{timestamp}.xlsx"

    # Save detailed file
    combined_df.to_excel(detailed_path, index=False)

    # Save summary file with chart using XlsxWriter
    with pd.ExcelWriter(summary_path, engine='xlsxwriter') as writer:
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
        workbook = writer.book
        worksheet = writer.sheets['Summary']

        # Create a chart
        chart = workbook.add_chart({'type': 'column'})

        chart.add_series({
            'name': 'Quality %',
            'categories': ['Summary', 1, 0, len(summary_df), 0],  # User Name
            'values': ['Summary', 1, 3, len(summary_df), 3],      # Quality %
            'fill': {'color': '#4F81BD'},

        })

        chart.set_title({'name': 'User-wise Quality %'})
        chart.set_x_axis({'name': 'User Name'})
        chart.set_y_axis({'name': 'Quality %', 'major_gridlines': {'visible': False}})
        chart.set_style(11)

        # Insert chart
        worksheet.insert_chart('F2', chart, {'x_scale': 1.3, 'y_scale': 1.2})


    print(f"\n📤 Reports generated:")
    print("🔹 Summary Report:", summary_path)
    print("🔹 Detailed Report:", detailed_path)

    return summary_path, detailed_path, graph_data




@app.route('/generate_project_report', methods=['POST'])
def generate_project_report():
    selected_files = request.form.getlist('selected_files')

    # ✅ Debugging output to console
    print("Selected files:", selected_files)

    if not selected_files:
        flash("No files selected.")
        return redirect(url_for('existing_projects'))

    summary_path, detailed_path,graph_data = process_quality_reports(selected_files)

    if summary_path and detailed_path:
        return render_template(
            'quality_report_result.html',
            summary_file=summary_path,
            detailed_file=detailed_path,
            graph_data=graph_data
        )

    else:
        flash("No valid reports to process.")
        return redirect(url_for('existing_projects'))


@app.route('/delete_file/<int:file_id>', methods=['POST'])
def delete_file(file_id):
    cur = mysql.connection.cursor()

    # Get the file path
    cur.execute("SELECT filepath FROM project_files WHERE pro_id = %s", (file_id,))
    file_data = cur.fetchone()
    if not file_data:
        flash("File not found in database", "error")
        return redirect(url_for('existing_projects'))

    filepath = file_data[0]

    # Delete the file from disk
    if filepath and os.path.exists(filepath):
        os.remove(filepath)

    # Delete from database
    cur.execute("DELETE FROM project_files WHERE pro_id = %s", (file_id,))
    mysql.connection.commit()
    cur.close()

    flash("File deleted successfully", "success")
    return redirect(url_for('existing_projects'))


@app.route('/delete_project/<int:project_id>', methods=['POST'])
def delete_project(project_id):
    cursor = mysql.connection.cursor()

    # Step 1: Delete all files (optional: if stored in DB or filesystem)
    cursor.execute("DELETE FROM project_files WHERE project_id = %s", (project_id,))

    # Step 2: Delete all members linked to this project
    cursor.execute("DELETE FROM members WHERE project_id = %s", (project_id,))

    # Step 3: Delete the project itself
    cursor.execute("DELETE FROM projects WHERE project_id = %s", (project_id,))

    mysql.connection.commit()
    cursor.close()

    flash("✅ Project deleted successfully!", "popup")


    return redirect(url_for('existing_projects'))


@app.route('/edit_project/<int:project_id>', methods=['GET', 'POST'])
def edit_project(project_id):
    cursor = mysql.connection.cursor()

    if request.method == 'POST':
        # Update existing members
        member_ids = request.form.getlist('member_id')
        names = request.form.getlist('name')
        emails = request.form.getlist('email')
        phones = request.form.getlist('phone')

        for i in range(len(names)):
            if member_ids[i]:  # Existing member
                cursor.execute("""
                    UPDATE members SET name=%s, email=%s, phone=%s
                    WHERE id=%s AND project_id=%s
                """, (names[i], emails[i], phones[i], member_ids[i], project_id))
            else:  # New member
                cursor.execute("""
                    INSERT INTO members (project_id, name, email, phone)
                    VALUES (%s, %s, %s, %s)
                """, (project_id, names[i], emails[i], phones[i]))

        mysql.connection.commit()
        cursor.close()
        flash("Team members updated successfully", "success")
        return redirect(url_for('existing_projects'))

    # GET: Fetch existing members
    cursor.execute("SELECT * FROM members WHERE project_id = %s", (project_id,))
    members = cursor.fetchall()
    cursor.close()
    return render_template('edit_project.html', project_id=project_id, members=members)
@app.route('/delete_member/<int:member_id>/<int:project_id>', methods=['POST'])
def delete_member(member_id, project_id):
    cursor = mysql.connection.cursor()
    cursor.execute("DELETE FROM members WHERE id = %s", (member_id,))
    mysql.connection.commit()
    cursor.close()
    flash("Member deleted successfully", "success")
    return redirect(url_for('edit_project', project_id=project_id))



@app.route('/clear_session')
def clear_session():
    session.clear()
    return "Session cleared"

@app.route('/logout')
def logout():
    session.clear()
    flash("🚪 Logout successful.", "success")
    return redirect(url_for('start_page'))


def is_valid_softnis(series):
    pattern = re.compile(r'^(?=.*[A-Za-z])[A-Za-z0-9_]+$')
    for val in series:
        val = str(val).strip()
        if not val:
            continue
        if not pattern.match(val):
            return False
    return True

@app.route('/send_to_members/<int:project_id>/<filename>', methods=['GET'])
def show_members(project_id, filename):
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT id, name, email FROM members WHERE project_id = %s", (project_id,))
    members = cursor.fetchall()
    return render_template('send_to_members.html', members=members, project_id=project_id, filename=filename)
@app.route('/send_excel_to_selected_members', methods=['POST'])
def send_excel_to_selected_members():
    selected_emails = request.form.getlist('selected_members')
    filename = request.form['filename']
    project_id = request.form['project_id']

    # Build the full file path
    file_path = os.path.join(app.root_path, 'static', 'results', filename)

    # Email content
    subject = "Project Report"
    body = "Please find the attached Excel report for your project."

    try:
        for email in selected_emails:
            msg = Message(subject, recipients=[email])
            msg.body = body

            with app.open_resource(file_path) as fp:
                msg.attach(
                    filename,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    fp.read()
                )

            mail.send(msg)

        flash("📤 Successfully delivered the report via email..", "success")

    except Exception as e:
        flash(f"❌ Failed to send email: {str(e)}", "danger")

    # Redirect back to result page
    download_link = url_for('download_file', filename=filename)
    return render_template('result.html', filename=filename, project_id=project_id, download_link=download_link)



# ✅ CHANGED ROUTE FROM '/' TO '/upload'
@app.route('/upload', methods=['GET', 'POST'])
def index():
    if not session.get('logged_in'):
        flash(" Please authenticate to access this resource.", "popup")
        return redirect('/login')

    # ✅ Fetch all existing projects to display in dropdown
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT project_id,name FROM projects")
    projects = cursor.fetchall()

    if request.method == 'POST':
        project_id = request.form.get('project_id')  # 🔸 Get selected project ID
        file = request.files.get('file')

        if not project_id:
            flash("⚠️ Please select a Project ID before uploading.", "popup")
            return render_template('index.html', projects=projects)

        if not file or not file.filename.endswith(('.xls', '.xlsx')):
            flash("Please upload a valid Excel file with .xls or .xlsx extension.", "popup")
            return render_template('index.html', projects=projects)

        filename = file.filename
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)

        try:
            excel_file = pd.ExcelFile(file_path)
            sheets = excel_file.sheet_names

            if len(sheets) != 2:
                flash("Uploaded file must have exactly 2 sheets.", "popup")
                return render_template('index.html', projects=projects)

            if 'Production Completed' not in sheets or 'Delivered' not in sheets:
                flash("Sheet names must be exactly 'Production Completed' and 'Delivered'.", "popup")
                return render_template('index.html', projects=projects)

            df_prod = pd.read_excel(file_path, sheet_name='Production Completed')
            df_del = pd.read_excel(file_path, sheet_name='Delivered')

            if 'User Name' not in df_prod.columns or 'User Name' not in df_del.columns:
                flash("Unexpected error: 'User Name' column missing.", "popup")
                return render_template('index.html', projects=projects)

            prod_id_col = next((col for col in df_prod.columns if str(col).strip().lower() == "softnis id"), None)
            del_id_col = next((col for col in df_del.columns if str(col).strip().lower() == "softnis id"), None)

            if not prod_id_col or not del_id_col:
                flash("'SoftNis ID' column not found in one or both sheets.", "popup")
                return render_template('index.html', projects=projects)

            if not is_valid_softnis(df_prod[prod_id_col]) or not is_valid_softnis(df_del[del_id_col]):
                flash("Some 'SoftNis ID's are invalid. Only letters, numbers, and underscores are allowed.", "popup")
                return render_template('index.html', projects=projects)

            result_filename = f"result_{filename}"
            results_dir = os.path.join(app.root_path, 'static', 'results')
            os.makedirs(results_dir, exist_ok=True)
            result_path = os.path.join(results_dir, result_filename)


            # 🔸 Generate comparison report
            generate_report(file_path, df_prod, df_del, result_path)

            # 🔸 Save the result file in project_files table
            cursor.execute("""
                INSERT INTO project_files (project_id, filename, filepath)
                VALUES (%s, %s, %s)
            """, (project_id, result_filename, result_path))
            mysql.connection.commit()
            cursor.close()

            flash("✅ File processed successfully and saved under the selected project.", "success")
            return render_template(
                'result.html',
                download_link=url_for('download_file', filename=result_filename),
                filename=result_filename,
                project_id=project_id  # ✅ this fixes the undefined error
            )

        except Exception as e:
            import traceback
            print("DEBUG ERROR:\n", traceback.format_exc())
            flash(f"❌ Unexpected error occurred: {str(e)}", "popup")
            return render_template('index.html', projects=projects)

    return render_template('index.html', projects=projects)


@app.route('/assign_project', methods=['GET', 'POST'])
def assign_project():
    filename = request.args.get('filename')  # comes from previous redirect
    filepath = os.path.join(RESULT_FOLDER, filename)

    if request.method == 'POST':
        project_id = request.form.get('project_id')

        cursor = mysql.connection.cursor()
        cursor.execute("""
            INSERT INTO project_files (project_id, filename, filepath)
            VALUES (%s, %s, %s)
        """, (project_id, filename, filepath))
        mysql.connection.commit()
        cursor.close()

        flash("✅ File successfully linked to the selected project.", "success")
        return redirect(url_for('existing_projects'))

    # On GET: Fetch all existing projects for dropdown
    cursor = mysql.connection.cursor()
    cursor.execute("SELECT project_id, project_name FROM projects")
    projects = cursor.fetchall()
    cursor.close()

    return render_template('assign_project.html', filename=filename, projects=projects)


@app.route('/download/<filename>')
def download_file(filename):
    results_dir = os.path.join(app.root_path, 'static', 'results')
    full_path = os.path.join(results_dir, filename)

    if os.path.exists(full_path):
        return send_file(full_path, as_attachment=True)
    else:
        flash("File not found", "popup")
        return redirect(url_for('existing_projects'))


def generate_report(file_path, df_prod, df_del, result_path): # main logic
    # Normalize and match all Attribute Name variants
    ignore_cols = [col.lower().strip() for col in df_prod.columns if 'attribute name' in col.lower()]
    ignore_cols += ['softnis id', 'user name']  # Add others to skip
    # Normalize Delivered sheet column names to match Production sheet
    for i in range(56):  # Assume up to 56 attribute columns fix coloumn mismatch
        name_col = f"Technical Specification {i+1} Name"
        value_col = f"Technical Specification {i+1} Value"
        attr_name = f"Attribute Name.{i}" if i != 0 else "Attribute Name"
        attr_value = f"Attribute Value.{i}" if i != 0 else "Attribute Value"
        if name_col in df_del.columns:
            df_del.rename(columns={name_col: attr_name}, inplace=True)
        if value_col in df_del.columns:
            df_del.rename(columns={value_col: attr_value}, inplace=True)


    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Ensure 'SoftNis ID' is present finds  tht column tht holds softnis id in df_prod
    id_col = next((col for col in df_prod.columns if str(col).strip().lower() == "softnis id"), None)
    if id_col is None:
        raise ValueError("SoftNis ID column not found in Production Completed sheet.")
     #converts delivered dataframe (df_del) into a dictionary converts index dataframe into nested dictionary
    df_del = df_del.drop_duplicates(subset=[id_col]) # prepare delivered sheet for comparison
    df_del_dict = df_del.set_index(id_col).to_dict('index') # create a dictionary of rows indexed by softnis id for fast lookup
#Uses openpyxl to load the original Excel workbook (file_path is the uploaded Excel file).
    wb = load_workbook(file_path)    # loads the workbook and selects the worksheet for editing
    ws_prod = wb["Production Completed"]
    headers = list(df_prod.columns)  # retrieve column header
    user_col_index = next((i for i, col in enumerate(headers) if str(col).strip().lower() == 'user name'), -1) # loops throught all column names and check if any of them is uername is case insensitive,stripped of extra spaces
    if user_col_index == -1:
        raise ValueError("User Name column not found.")

    quality_col_start = len(headers) + 1
    ws_prod.cell(row=1, column=quality_col_start).value = "Row Quality %"
    ws_prod.cell(row=1, column=quality_col_start + 1).value = "Right Values"
    ws_prod.cell(row=1, column=quality_col_start + 2).value = "Wrong Values"
    ws_prod.cell(row=1, column=quality_col_start + 3).value = "Error Report"

    user_stats = {}

    for i in range(len(df_prod)): # row by row comparison
        right = 0
        wrong = 0

        prod_row = df_prod.iloc[i]
        softnis_id = str(prod_row[id_col]).strip()
        if not softnis_id or softnis_id not in df_del_dict:
            # Write error in "Error Report" column
            ws_prod.cell(row=i + 2, column=quality_col_start + 3).value = "SoftNis ID not found in Delivered"
            # Highlight full row
            total_cols = quality_col_start + 3
            for col_index in range(1, total_cols + 1):
                ws_prod.cell(row=i + 2, column=col_index).fill = red_fill
            continue

        del_row = df_del_dict[softnis_id]

        for j, col in enumerate(headers):
            col_name = str(col).strip().lower()
            if col_name in ignore_cols:
                continue

            val1 = str(prod_row[col]).strip() if pd.notna(prod_row[col]) else ""
            val2 = str(del_row.get(col, "")).strip() if pd.notna(del_row.get(col, "")) else ""

            if not val1 and not val2:
                continue

            if val1 == val2 and val1 != "":
                right += 1
            else:
                if val1 != "" or val2 != "":
                    ws_prod.cell(row=i + 2, column=j + 1).fill = red_fill
                    wrong += 1

        total = right + wrong
        quality = (right / total * 100) if total > 0 else 0

        ws_prod.cell(row=i + 2, column=quality_col_start).value = round(quality, 2)
        ws_prod.cell(row=i + 2, column=quality_col_start + 1).value = right
        ws_prod.cell(row=i + 2, column=quality_col_start + 2).value = wrong
        # ✅ Check User Name mismatch
        # Compare User Name for mismatch
        prod_user = str(prod_row.get("User Name", "")).strip().lower()
        del_user = str(del_row.get("User Name", "")).strip().lower()

        if prod_user and del_user and prod_user != del_user:
            error_cell = ws_prod.cell(row=i + 2, column=quality_col_start + 3)
            error_cell.value = "User Name mismatch"
            error_cell.fill = red_fill

            # ✅ Fill the whole row red (including all original columns + quality columns + error report)
            total_cols = quality_col_start + 3  # includes error report
            for col_index in range(1, total_cols + 1):
                ws_prod.cell(row=i + 2, column=col_index).fill = red_fill

        username = str(prod_row[user_col_index]).strip()
        if username:
            if username not in user_stats:
                user_stats[username] = {'correct': right, 'total': total}
            else:
                user_stats[username]['correct'] += right
                user_stats[username]['total'] += total

    # Write Quality Report
    if "Quality Report" in wb.sheetnames:
        wb.remove(wb["Quality Report"])
    ws_report = wb.create_sheet("Quality Report")
    ws_report.append(["User Name", "Matched Cells", "Total Cells", "Quality %"])

    # After appending user stats rows to "Quality Report" sheet
    for user, stats in user_stats.items():
        q = (stats['correct'] / stats['total'] * 100) if stats['total'] else 0
        ws_report.append([user, stats['correct'], stats['total'], round(q, 2)])

    # ✅ Create a bar chart to visualize Quality % by User Name
    from openpyxl.chart import BarChart, Reference

    chart = BarChart()
    chart.title = "User Quality %"
    chart.y_axis.title = "Quality Percentage"
    chart.x_axis.title = "User Name"

    # Use the "Quality %" column (4th column), from row 1 to last
    data = Reference(ws_report, min_col=4, min_row=1, max_row=ws_report.max_row)

    # X-axis: User names (1st column), from row 2 to last
    categories = Reference(ws_report, min_col=1, min_row=2, max_row=ws_report.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Optional: size of the chart
    chart.width = 20
    chart.height = 10

    # Add chart to the sheet starting at cell E2
    ws_report.add_chart(chart, "E2")

 # ✅ Don't forget to save the workbook
    print("Saving to:", result_path)
    wb.save(result_path)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
