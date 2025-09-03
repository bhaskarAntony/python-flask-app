from flask import Flask, request, render_template, send_file, url_for, flash, session, redirect
import pandas as pd #read and manipulate
import os  # file handling
from openpyxl import load_workbook  # modify excel sheets
from openpyxl.cell import cell
from openpyxl.styles import PatternFill  #  patternFill add red color fill for missmatched cells
import re
app = Flask(__name__)
app.secret_key = 'your_super_secret_key'  # 🛡️ Required for session and flash
from flask_mail import Mail, Message
import random

# Add mail config (use your mail server here)
app.config.update(
    MAIL_SERVER='smtp.gmail.com',
    MAIL_PORT=587,
    MAIL_USE_TLS=True,
    MAIL_USERNAME='reshmamohan938@gmail.com',  # ✅ Your Gmail address
    MAIL_PASSWORD='hqxr hngg dtjw gzcx'      # ✅ App Password, NOT your Gmail login
)


mail = Mail(app)


UPLOAD_FOLDER = 'uploads'  # create folder to store uploaded files and result files
RESULT_FOLDER = 'results'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True) # ensure those folder exist
import re  # already correctly placed

@app.route('/login', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    allowed_domains = ['@softnis.com']
    allowed_emails = ['softnisdata@gmail.com', 'reshmamohan938@gmail.com']

    if request.method == 'POST':
        email = request.form['email']

        # ✅ Check email domain or exact match
        if not any(email.endswith(domain) for domain in allowed_domains) and email not in allowed_emails:
            flash("❌ Please use your company email (like name@softnis.com)", "popup")
            return render_template('login.html')  # ✅ Return here on error

        # ✅ Generate and send OTP
        otp = str(random.randint(1000, 9999))
        session['otp'] = otp
        session['email'] = email

        msg = Message('Your OTP for Login', sender='your_email@gmail.com', recipients=[email])
        msg.body = f'Your OTP is: {otp}'
        mail.send(msg)

        flash("✅ OTP sent to your email.", "success")
        return redirect(url_for('verify'))  # ✅ Successful flow

    return render_template('login.html')  # ✅ This covers GET request



@app.route('/verify', methods=['GET', 'POST'])
def verify():
    if request.method == 'POST':
        entered_otp = request.form.get('otp')
        if entered_otp == session.get('otp'):
            session['logged_in'] = True
            flash("✅ Login successful!", "success")
            return redirect('/')
        else:
            flash("❌ Incorrect OTP. Please try again.", "popup")
            return render_template('verify.html')

    return render_template('verify.html')
@app.route('/logout')
def logout():
    session.clear()  # Clears all session data (including 'logged_in', 'email', 'otp')
    flash("🚪 You have been logged out successfully.", "success")
    return redirect(url_for('login'))


# ✅ PLACE THIS FUNCTION HERE
def is_valid_softnis(series):
    pattern = re.compile(r'^(?=.*[A-Za-z])[A-Za-z0-9_]+$')  # allows alphanumeric + underscore; must contain at least one letter
    for val in series:
        val = str(val).strip()
        if not val:  # skip blanks
            continue
        if not pattern.match(val):
            return False
    return True

@app.route('/', methods=['GET', 'POST'])   # get= shows upload page and post=process the uploaded file
def index():
    if not session.get('logged_in'):
        flash("⚠️ Please log in to access this page.", "popup")
        return redirect('/login')
    if request.method == 'POST':
        file = request.files.get('file')    # get the uploaded file
        if not file or not file.filename.endswith(('.xls', '.xlsx')):
            flash("❌ Please upload a valid Excel file with .xls or .xlsx extension.", "popup")# reject the non excel files
            return "Please upload a valid Excel file."

        filename = file.filename
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(file_path)   # save a file to the desk

        try:
            # Try loading Excel file
            excel_file = pd.ExcelFile(file_path)
            sheets = excel_file.sheet_names

            if len(sheets) != 2:
                flash("❌ Uploaded file must have exactly 2 sheets.", "popup")
                return render_template('index.html')

            if 'Production Completed' not in sheets or 'Delivered' not in sheets:
                flash("❌ Sheet names must be exactly 'Production Completed' and 'Delivered'.", "popup")
                return render_template('index.html')

            # ✅ Load data and run comparison
            df_prod = pd.read_excel(file_path, sheet_name='Production Completed')
            df_del = pd.read_excel(file_path, sheet_name='Delivered')
            # ✅ Check if 'User Name' column exists in both sheets
            if 'User Name' not in df_prod.columns or 'User Name' not in df_del.columns:
                flash("❌ Unexpected error occurred: 'User Name' column not found in one or both sheets.", "popup")
                return render_template('index.html')

            # Identify SoftNis ID columns
            # Identify SoftNis ID columns
            prod_id_col = next((col for col in df_prod.columns if str(col).strip().lower() == "softnis id"), None)
            del_id_col = next((col for col in df_del.columns if str(col).strip().lower() == "softnis id"), None)

            if not prod_id_col or not del_id_col:
                flash("❌ 'SoftNis ID' column not found in one or both sheets.", "popup")
                return render_template('index.html')

            # ✅ Allow alphanumeric + underscore IDs like HCPB11_1


            if not is_valid_softnis(df_prod[prod_id_col]):
                flash(
                    "❌ Some 'SoftNis ID's in 'Production Completed' are invalid. Only letters, numbers, and underscores are allowed.",
                    "popup")
                return render_template('index.html')

            if not is_valid_softnis(df_del[del_id_col]):
                flash(
                    "❌ Some 'SoftNis ID's in 'Delivered' are invalid. Only letters, numbers, and underscores are allowed.",
                    "popup")
                return render_template('index.html')

            result_filename = f"result_{filename}"
            result_path = os.path.join(RESULT_FOLDER, result_filename)

            generate_report(file_path, df_prod, df_del, result_path)

            flash("✅ File processed successfully. Ready to download.", "success")
            return render_template('result.html', download_link=url_for('download_file', filename=result_filename))


        except Exception as e:
            import traceback
            error_msg = traceback.format_exc()
            print("DEBUG ERROR:\n", error_msg)  # Logs full traceback in terminal
            flash(f"❌ Unexpected error occurred: {str(e)}", "popup")
            return render_template('index.html')


    return render_template('index.html')

@app.route('/download/<filename>')  # download route
def download_file(filename):
    return send_file(os.path.join(RESULT_FOLDER, filename), as_attachment=True)  # sends the generated report to the browser for download

def generate_report(file_path, df_prod, df_del, result_path): # main logic
    # Normalize and match all Attribute Name variants
    ignore_cols = [col.lower().strip() for col in df_prod.columns if 'attribute name' in col.lower()]
    ignore_cols += ['softnis id', 'user name']  # Add others to skip
    # Normalize Delivered sheet column names to match Production sheet
    for i in range(56):  # Assume up to 56 attribute columns fix coloumn mismatch
        name_col = f"Technical Specification {i+1} Name"
        value_col = f"Technical Specification {i+1} Value"
        attr_name = f"Attribute Name.{i}" if i != 0 else "Attribute Name"
        attr_value = f"Attribute Value.{i}" if i != 0 else "Attribute Value"
        if name_col in df_del.columns:
            df_del.rename(columns={name_col: attr_name}, inplace=True)
        if value_col in df_del.columns:
            df_del.rename(columns={value_col: attr_value}, inplace=True)


    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Ensure 'SoftNis ID' is present finds  tht column tht holds softnis id in df_prod
    id_col = next((col for col in df_prod.columns if str(col).strip().lower() == "softnis id"), None)
    if id_col is None:
        raise ValueError("SoftNis ID column not found in Production Completed sheet.")
     #converts delivered dataframe (df_del) into a dictionary converts index dataframe into nested dictionary
    df_del = df_del.drop_duplicates(subset=[id_col]) # prepare delivered sheet for comparison
    df_del_dict = df_del.set_index(id_col).to_dict('index') # create a dictionary of rows indexed by softnis id for fast lookup
#Uses openpyxl to load the original Excel workbook (file_path is the uploaded Excel file).
    wb = load_workbook(file_path)    # loads the workbook and selects the worksheet for editing
    ws_prod = wb["Production Completed"]
    headers = list(df_prod.columns)  # retrieve column header
    user_col_index = next((i for i, col in enumerate(headers) if str(col).strip().lower() == 'user name'), -1) # loops throught all column names and check if any of them is uername is case insensitive,stripped of extra spaces
    if user_col_index == -1:
        raise ValueError("User Name column not found.")

    quality_col_start = len(headers) + 1
    ws_prod.cell(row=1, column=quality_col_start).value = "Row Quality %"
    ws_prod.cell(row=1, column=quality_col_start + 1).value = "Right Values"
    ws_prod.cell(row=1, column=quality_col_start + 2).value = "Wrong Values"
    ws_prod.cell(row=1, column=quality_col_start + 3).value = "Error Report"

    user_stats = {}

    for i in range(len(df_prod)): # row by row comparison
        right = 0
        wrong = 0

        prod_row = df_prod.iloc[i]
        softnis_id = str(prod_row[id_col]).strip()
        if not softnis_id or softnis_id not in df_del_dict:
            # Write error in "Error Report" column
            ws_prod.cell(row=i + 2, column=quality_col_start + 3).value = "SoftNis ID not found in Delivered"
            # Highlight full row
            total_cols = quality_col_start + 3
            for col_index in range(1, total_cols + 1):
                ws_prod.cell(row=i + 2, column=col_index).fill = red_fill
            continue

        del_row = df_del_dict[softnis_id]

        for j, col in enumerate(headers):
            col_name = str(col).strip().lower()
            if col_name in ignore_cols:
                continue

            val1 = str(prod_row[col]).strip() if pd.notna(prod_row[col]) else ""
            val2 = str(del_row.get(col, "")).strip() if pd.notna(del_row.get(col, "")) else ""

            if not val1 and not val2:
                continue

            if val1 == val2 and val1 != "":
                right += 1
            else:
                if val1 != "" or val2 != "":
                    ws_prod.cell(row=i + 2, column=j + 1).fill = red_fill
                    wrong += 1

        total = right + wrong
        quality = (right / total * 100) if total > 0 else 0

        ws_prod.cell(row=i + 2, column=quality_col_start).value = round(quality, 2)
        ws_prod.cell(row=i + 2, column=quality_col_start + 1).value = right
        ws_prod.cell(row=i + 2, column=quality_col_start + 2).value = wrong
        # ✅ Check User Name mismatch
        # Compare User Name for mismatch
        prod_user = str(prod_row.get("User Name", "")).strip().lower()
        del_user = str(del_row.get("User Name", "")).strip().lower()

        if prod_user and del_user and prod_user != del_user:
            error_cell = ws_prod.cell(row=i + 2, column=quality_col_start + 3)
            error_cell.value = "User Name mismatch"
            error_cell.fill = red_fill

            # ✅ Fill the whole row red (including all original columns + quality columns + error report)
            total_cols = quality_col_start + 3  # includes error report
            for col_index in range(1, total_cols + 1):
                ws_prod.cell(row=i + 2, column=col_index).fill = red_fill

        username = str(prod_row[user_col_index]).strip()
        if username:
            if username not in user_stats:
                user_stats[username] = {'correct': right, 'total': total}
            else:
                user_stats[username]['correct'] += right
                user_stats[username]['total'] += total

    # Write Quality Report
    if "Quality Report" in wb.sheetnames:
        wb.remove(wb["Quality Report"])
    ws_report = wb.create_sheet("Quality Report")
    ws_report.append(["User Name", "Matched Cells", "Total Cells", "Quality %"])

    # After appending user stats rows to "Quality Report" sheet
    for user, stats in user_stats.items():
        q = (stats['correct'] / stats['total'] * 100) if stats['total'] else 0
        ws_report.append([user, stats['correct'], stats['total'], round(q, 2)])

    # ✅ Create a bar chart to visualize Quality % by User Name
    from openpyxl.chart import BarChart, Reference

    chart = BarChart()
    chart.title = "User Quality %"
    chart.y_axis.title = "Quality Percentage"
    chart.x_axis.title = "User Name"

    # Use the "Quality %" column (4th column), from row 1 to last
    data = Reference(ws_report, min_col=4, min_row=1, max_row=ws_report.max_row)

    # X-axis: User names (1st column), from row 2 to last
    categories = Reference(ws_report, min_col=1, min_row=2, max_row=ws_report.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    # Optional: size of the chart
    chart.width = 20
    chart.height = 10

    # Add chart to the sheet starting at cell E2
    ws_report.add_chart(chart, "E2")


if __name__ == '__main__':
    app.run(debug=True)
